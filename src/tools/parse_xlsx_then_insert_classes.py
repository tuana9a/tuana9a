import json
import yaml
import requests
import openpyxl
import base64
import time


class SchoolClass:
    def __init__(self, row, prop_table, semester):
        self.prop_table = prop_table
        self.semester = semester
        for key in prop_table:
            self.__setattr__(key, row[prop_table[key]])
        self.created = int(time.time() * 1000)
        self.semester = semester
        pass

    def __str__(self) -> str:
        result = ""
        for key in self.prop_table:
            result = result + key + ": " + \
                str(self.__getattribute__(key)) + "|"
        result += "semester: " + str(self.semester) + "|"
        result += "created: " + str(self.created)
        return result

    def to_dict(self):
        result = {}
        for key in self.prop_table:
            result[key] = self.__getattribute__(key)
        result["created"] = self.created
        result["semester"] = self.semester
        return result


CONFIG = yaml.load(open('tools.config.local.yaml'), Loader=yaml.FullLoader)

SECRET = CONFIG['secret']
print(f"secret=${SECRET}")

ENDPOINT = CONFIG['endpoint']
print(f"endpoint={ENDPOINT}")

ENTRIES = CONFIG['entries']

COLUMN_NAME_TO_PROP_MAPPER = {
    "Mã_lớp": "MaLop",
    "Mã_lớp_kèm": "MaLopKem",
    "Mã_HP": "MaHocPhan",
    "Tên_HP": "TenHocPhan",
    "Buổi_số": "BuoiHocSo",
    "Thứ": "HocVaoThu",
    "Phòng": "PhongHoc",
    "Thời_gian": "ThoiGianHoc",
    "Tuần": "HocVaoCacTuan",
    "Loại_lớp": "LoaiLop",
    "Ghi_chú": "GhiChu",
}


def run(semester, xlsx_file_path):
    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active
    max_column = ws.max_column
    max_row = ws.max_row
    handler_name = "not"
    current_row = 0
    classes_to_insert = []
    classes_count = 0
    prop_table = {
        "MaLop": -1,
        "MaLopKem": -1,
        "MaHocPhan": -1,
        "TenHocPhan": -1,
        "BuoiHocSo": -1,
        "HocVaoThu": -1,
        "PhongHoc": -1,
        "ThoiGianHoc": -1,
        "HocVaoCacTuan": -1,
        "LoaiLop": -1,
        "GhiChu": -1,
    }

    print(f"semester={semester}")
    print(f"xlsx_path=${xlsx_file_path}")
    print(f"max_column={max_column}")
    print(f"max_row={max_row}")

    def check_found_prop_table():
        for key in prop_table:
            if prop_table[key] != -1:
                return True
        return False

    def find_prop_indexes(row):
        i = 0
        for cell in row:
            prop_name = COLUMN_NAME_TO_PROP_MAPPER.get(cell)
            if prop_name:
                prop_table[prop_name] = i
            i = i + 1

    def handle_row(row):
        school_class = SchoolClass(row, prop_table, semester)
        classes_to_insert.append(school_class.to_dict())

    handlers = {
        "not": find_prop_indexes,
        "found": handle_row,
    }

    for row in ws.iter_rows(min_row=0, values_only=True):
        current_row += 1
        handlers.get(handler_name)(row)
        if check_found_prop_table():
            handler_name = "found"

    if len(classes_to_insert) > 0:
        headers = {'Authorization': f'Basic {base64.b64encode(SECRET.encode("utf-8")).decode("utf-8")}'}
        batch = []
        max_batch_size = 100
        for classs in classes_to_insert:
            if len(batch) > max_batch_size:
                response = requests.post(
                    url=ENDPOINT, headers=headers, json={'classes': batch})
                batch.clear()
                print(f'response.status={response.status_code}',
                      f'response.text={response.text}')
                classes_count += int(json.loads(response.text)["data"])
            else:
                batch.append(classs)
        print('count =', classes_count)


for entry in ENTRIES:
    run(entry['semester'], entry['file_path'])
