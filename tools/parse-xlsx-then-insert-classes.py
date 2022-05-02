import os
import yaml
import requests
import openpyxl
from data import SchoolClass

CONFIG = yaml.load(open('tools.config.local.yaml'), Loader=yaml.FullLoader)

SECRET = CONFIG['secret']
print(f"secret=${SECRET}")

ENDPOINT = CONFIG['endpoint']
print(f"endpoint={ENDPOINT}")

ENTRIES = CONFIG['entries']

COLUMN_NAME_TO_PROP_MAPPER = {
    "Mã_lớp": "MaLop",
    "Mã_lớp_kèm": "MaLopKem",
    "Mã_HP": "MaHocPhan",
    "Tên_HP": "TenHocPhan",
    "Buổi_số": "BuoiHocSo",
    "Thứ": "HocVaoThu",
    "Phòng": "PhongHoc",
    "Thời_gian": "ThoiGianHoc",
    "Tuần": "HocVaoCacTuan",
    "Loại_lớp": "LoaiLop",
    "Ghi_chú": "GhiChu",
}

def run(semester, xlsx_file_path):
    print(f"semester={semester}")
    print(f"xlsx_path=${xlsx_file_path}")

    prop_table = {
        "MaLop": -1,
        "MaLopKem": -1,
        "MaHocPhan": -1,
        "TenHocPhan": -1,
        "BuoiHocSo": -1,
        "HocVaoThu": -1,
        "PhongHoc": -1,
        "ThoiGianHoc": -1,
        "HocVaoCacTuan": -1,
        "LoaiLop": -1,
        "GhiChu": -1,
    }

    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active
    max_column = ws.max_column
    max_row = ws.max_row

    print(f"max_column={max_column}")
    print(f"max_row={max_row}")

    found_prop_table = False
    prompt_to_process = False
    current_row = 0

    def check_found_prop_table():
        for key in prop_table:
            if prop_table[key] != -1:
                return True
        return False

    classes_to_insert = []
    classes_count = 0

    for row in ws.iter_rows(min_row=0, values_only=True):
        current_row += 1
        if not found_prop_table:
            # determine column name then map it to prop table
            i = 0
            for cell in row:
                prop_name = COLUMN_NAME_TO_PROP_MAPPER.get(cell)
                if(prop_name):
                    prop_table[prop_name] = i
                i = i + 1
            if check_found_prop_table():
                found_prop_table = True
                print("\n*** prop_table ***")
                for key in prop_table:
                    print(key, "=", prop_table[key], end="|")
                value = input("\nis that ok ? (Y) to continue (Other) to cancel: ")
                if value == "Y":
                    prompt_to_process = True
                    print(f'start_data_row = {current_row}')
        elif (prompt_to_process):
            classes_count = classes_count + 1
            school_class = SchoolClass(row, prop_table, semester)
            classes_to_insert.append(school_class.to_dict())

    if len(classes_to_insert) > 0:
        headers = { 'secret': SECRET }
        batch = []
        max_batch_size = 50
        for classs in classes_to_insert:
            if len(batch) > max_batch_size:
                response = requests.post(url=ENDPOINT, headers=headers, json={ 'classes' : batch })
                batch.clear()
                print(f'response.status={response.status_code}', f'response.text={response.text}')
            else:
                batch.append(classs)
        print('count =', classes_count)

for entry in ENTRIES:
    run(entry['semester'], entry['file_path'])
